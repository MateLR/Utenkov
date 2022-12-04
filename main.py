import csv
import re
import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from jinja2 import Environment, FileSystemLoader
import prettytable
from prettytable import PrettyTable
import pdfkit


def exp_for_num(s: str):
    """Функция, котороая переводит строки с опытом в числа, которые в дальнейшем сравниваются

    Args:
        s (str): Принимает на вход одну переменную типа string, которая содержит в себе информацию об опыте работы

    Returns:
        int: возвращает числовое значение типа данных int, в соответствии с указанным значением опыта работы
    """
    s = re.findall(r'\d*\.\d+|\d+', s)
    return 0 if len(s) == 0 else int(s[0])


headings = ['№', 'Название', 'Описание', 'Навыки', 'Опыт работы', 'Премиум-вакансия', 'Компания',
            'Оклад', 'Название региона', 'Дата публикации вакансии']
currency = {"AZN": "Манаты",
            "BYR": "Белорусские рубли",
            "EUR": "Евро",
            "GEL": "Грузинский лари",
            "KGS": "Киргизский сом",
            "KZT": "Тенге",
            "RUR": "Рубли",
            "UAH": "Гривны",
            "USD": "Доллары",
            "UZS": "Узбекский сум"}
experience = {"noExperience": "Нет опыта",
              "between1And3": "От 1 года до 3 лет",
              "between3And6": "От 3 до 6 лет",
              "moreThan6": "Более 6 лет"}
bools = {"False": "Нет",
         "True": "Да"}
currency_to_rub = {
    "Манаты": 35.68,
    "Белорусские рубли": 23.91,
    "Евро": 59.90,
    "Грузинский лари": 21.74,
    "Киргизский сом": 0.76,
    "Тенге": 0.13,
    "Рубли": 1,
    "Гривны": 1.64,
    "Доллары": 60.66,
    "Узбекский сум": 0.0055,
}
functions_for_filter = {
    "Название": lambda vacancy, value: vacancy.name == value,
    "Описание": lambda vacancy, value: vacancy.description == value,
    "Компания": lambda vacancy, value: vacancy.employer_name == value,
    "Навыки": lambda vacancy, values: all(x in vacancy.key_skills for x in values.split(', ')),
    "Опыт работы": lambda vacancy, value: vacancy.experience_id == value,
    "Премиум-вакансия": lambda vacancy, value: vacancy.premium == value,
    "Название региона": lambda vacancy, value: vacancy.area_name == value,
    "Идентификатор валюты оклада": lambda vacancy, value: vacancy.salary.salary_currency == value,
    "Дата публикации вакансии": lambda vacancy, value: vacancy.published_at.strftime("%d.%m.%Y") == value,
    "Оклад": lambda vacancy, value: vacancy.salary.salary_from <= float(value) <= vacancy.salary.salary_to,
}
functions_for_sort = {
    "Название": lambda vacancy: vacancy.name,
    "Описание": lambda vacancy: vacancy.description,
    "Компания": lambda vacancy: vacancy.employer_name,
    "Навыки": lambda vacancy: len(vacancy.key_skills),
    "Опыт работы": lambda vacancy: exp_for_num(vacancy.experience_id),
    "Премиум-вакансия": lambda vacancy: vacancy.premium,
    "Название региона": lambda vacancy: vacancy.area_name,
    "Идентификатор валюты оклада": lambda vacancy: vacancy.salary.salary_currency,
    "Дата публикации вакансии": lambda vacancy: vacancy.published_at,
    "Оклад": lambda vacancy: vacancy.salary.mid_salary_in_rubles
}


class Salary:
    """Класс для представления зарплаты

    Attributes:
        salary_from (float): Нижняя граница вилки оклада
        salary_to (float): Верхняя граница вилки оклада
        salary_currency (str): Валюта оклада
        salary_gross (bool): Атрибут показывает есть ли налоговый вычет у зарплаты, по умолчанию значение False
        mid_salary_in_rubles (float): Среднее значение зарплаты в рублях
    """

    def __init__(self, salary: list):
        """Иницилизирует объект Salary, распаковывая все данные о зарплате, кроме налогового вычета

        Args:
            salary (list): Список данных о зарплате в порядке: Нижняя граница, Верхняя граница, Валюта
        """
        self.salary_from = float(salary[0])
        self.salary_to = float(salary[1])
        self.salary_currency = currency[salary[2]]
        self.salary_gross = False
        self.mid_salary_in_rubles = (self.salary_from + self.salary_to) / 2 * currency_to_rub[self.salary_currency]

    def add_gross(self, salary_gross: str):
        """Метод для изменения атрибута salary_gross, который нужно вызывать при наличии данной информации в строке

        Args:
            salary_gross (str): Строка с информацией на русском языке о наличии налогового вычета
        """
        self.salary_gross = bools[salary_gross]

    def to_string(self):
        """Преобразовывает всю информацию о зарплате в строку

        Returns (str): Информация о зарплате

        """
        return f'{"{:,d}".format(int(self.salary_from)).replace(",", " ")} - {"{:,d}".format(int(self.salary_to)).replace(",", " ")} ({self.salary_currency}) {"(Без вычета налогов)" if self.salary_gross != "Нет" else "(С вычетом налогов)"}'


class Vacancy(object):
    """Класс для представления Вакансии

    Attributes:
        name (str): Название вакансии
        salary (Salary): Вся информация о зарплате
        area_name (str): Название региона вакансии
        published_at (datetime): Дата публикации вакансии
        year (int): Год публикации вакансии
        description (str): Описание вакансии
        key_skills (list): Список навыков
        experience_id (str): Опыт работы требуемый для вакансии
        premium (bool): Примиальность вакансии
        employer_name (str): Название компании вакансии
    """

    def __init__(self, vacancy: dict):
        """Иницилизирует объект вакансии, распаковывает все данные и выполняет их конвертацию

        Args:
            vacancy (dict): Словарь с данными о вакансии
        """
        self.name = vacancy['name']
        self.salary = Salary(
            [vacancy['salary_from'], vacancy['salary_to'], vacancy['salary_currency']])
        self.area_name = vacancy['area_name']
        self.published_at = datetime.strptime(vacancy['published_at'], '%Y-%m-%dT%H:%M:%S%z')
        self.year = int(self.published_at.strftime("%Y"))
        if len(vacancy) > 6:
            self.description = vacancy['description']
            self.key_skills = vacancy['key_skills'].split(';;')
            self.experience_id = experience[vacancy['experience_id']]
            self.premium = bools[vacancy['premium']]
            self.employer_name = vacancy['employer_name']
            self.salary.add_gross(vacancy['salary_gross'])

    @staticmethod
    def cut_string(s: str):
        """Обрезает строку, если её длина больше 100 символов и добавляет многоточие в конце

        Args:
            s (str): Принимает на вход одну переменную типа string

        Returns:
            (str): Изменённая строка

        """
        if len(s) > 100:
            return s[:100] + '...'
        return s

    def get_row(self, number: int):
        """Преобразует всю информацию о вакансии в список для таблицы

        Args:
            number (int): Номер вакансии

        Returns:
            (list): Список данных о вакансии

        """
        return [number + 1, self.name, self.cut_string(self.description), self.cut_string('\n'.join(self.key_skills)),
                self.experience_id, self.premium, self.employer_name, self.salary.to_string(), self.area_name,
                self.published_at.strftime("%d.%m.%Y")]


class DataSet(object):
    """Класс, который преобразует csv файл в базу данных информации о вакансиях, и анализирует эту информацию

    Attributes:
        file_name (str):
        vacancies_objects (list): Список, хранящий вакансии в виде объекта Vacancy
        vacancies_number (int): Количество вакансий
        salary_by_years (dict): Словарь с зарплатами по годам
        number_by_years (dict): Словарь с количеством вакансий по годам
        salary_by_years_job (dict): Словарь с зарплатами по годам, по выбранной профессии
        number_by_years_job (dict): Словарь с количеством вакансий по годам, по выбранной профессии
        salary_by_area (dict): Словарь с зарплатами по регионам
        share_number_by_area (dict): Словарь с количеством зарплат по регионам
    """

    def __init__(self, file_name: str):
        """Инициализирует объект DataSet, преобразует файл с вакансиями в список вакансий

        Args:
            file_name: Имя файла
        """
        self.file_name = file_name
        self.vacancies_objects = [Vacancy(x) for x in self.file_to_rows()]
        self.vacancies_number = len(self.vacancies_objects)
        self.salary_by_years = dict()
        self.number_by_years = dict()
        self.salary_by_years_job = dict()
        self.number_by_years_job = dict()
        self.salary_by_area = dict()
        self.share_number_by_area = dict()

    def analyze(self, job_name: str):
        """Анализирует вакансии по названию профессии

        Args:
            job_name (str): Название профессии
        """
        self.fill_analyze_set(job_name)

        self.edit_analyze_set()

        self.print_analyze()

    def print_analyze(self):
        """Печатает в консоль данные с проведённого анализа вакансий
        """
        print(f"Динамика уровня зарплат по годам: {self.salary_by_years}")
        print(f"Динамика количества вакансий по годам: {self.number_by_years}")
        print(f"Динамика уровня зарплат по годам для выбранной профессии: {self.salary_by_years_job}")
        print(f"Динамика количества вакансий по годам для выбранной профессии: {self.number_by_years_job}")
        print(f"Уровень зарплат по городам (в порядке убывания): {self.salary_by_area}")
        print(f"Доля вакансий по городам (в порядке убывания): {self.share_number_by_area}")

    def fill_analyze_set(self, job_name: str):
        """Заполняет словари для анализа данными, которые потребуются для анализа

        Args:
            job_name (str): Название профессии
        """
        for vac in self.vacancies_objects:
            if vac.year not in self.number_by_years:
                self.number_by_years_job[vac.year] = 0
                self.salary_by_years_job[vac.year] = 0
                self.number_by_years[vac.year] = 0
                self.salary_by_years[vac.year] = 0
            if vac.area_name not in self.salary_by_area:
                self.share_number_by_area[vac.area_name] = 0
                self.salary_by_area[vac.area_name] = 0

            self.number_by_years[vac.year] = self.number_by_years[vac.year] + 1
            self.salary_by_years[vac.year] = self.salary_by_years[vac.year] + vac.salary.mid_salary_in_rubles
            self.share_number_by_area[vac.area_name] = self.share_number_by_area[vac.area_name] + 1
            self.salary_by_area[vac.area_name] = self.salary_by_area[vac.area_name] + vac.salary.mid_salary_in_rubles

            if vac.name.find(job_name) >= 0:
                self.number_by_years_job[vac.year] = self.number_by_years_job[vac.year] + 1
                self.salary_by_years_job[vac.year] = self.salary_by_years_job[
                                                         vac.year] + vac.salary.mid_salary_in_rubles

    def edit_analyze_set(self):
        """Редактирует словари для анализа данных, изменяя текущие данные под конечные, готовые к работе
        """
        for key in self.salary_by_years.keys():
            self.salary_by_years[key] = int(self.salary_by_years[key] / self.number_by_years[key]) if \
                self.number_by_years[key] != 0 else 0
        for key in self.salary_by_years_job.keys():
            self.salary_by_years_job[key] = int(self.salary_by_years_job[key] / self.number_by_years_job[key]) if \
                self.number_by_years_job[key] != 0 else 0

        areas = []
        for key in self.salary_by_area.keys():
            self.salary_by_area[key] = int(self.salary_by_area[key] / self.share_number_by_area[key])
            self.share_number_by_area[key] = round(self.share_number_by_area[key] / self.vacancies_number, 4)
            if self.share_number_by_area[key] < 0.01:
                areas.append(key)
        for key in areas:
            del self.salary_by_area[key]
            del self.share_number_by_area[key]

        self.salary_by_area = dict(sorted(self.salary_by_area.items(), key=lambda x: x[1], reverse=True)[:10])
        self.share_number_by_area = dict(
            sorted(self.share_number_by_area.items(), key=lambda x: x[1], reverse=True)[:10])

    @staticmethod
    def check_file_for_empty(len: int):
        """Проверяет входной файл на пустоту или отсутствия данных

        Args:
            len (int): Принимает на вход одну переменную типа int, длину массива данных из таблицы
        """
        if len < 2:
            print("Пустой файл" if len < 1 else "Нет данных")
            quit()

    @staticmethod
    def change_string(s: str):
        """Убирает html тэги в строке, лишние пробелы и заменяет переносы строки на специальную строку

        Args:
            s (str): Строка для обработки

        Returns:
            (str): Строка без html тэгов, лишних пробелов и переносов
        """
        s = s.replace('\n', ';;')
        return ' '.join(re.sub("<[^>]*>", "", s).split())

    def file_to_rows(self):
        """Извлекает данные из csv таблицы и преобразует их в список словарей, подходящих для преобразования в объект Vacancy

        Returns (list): Список словарей

        """
        r_file = open(self.file_name, encoding='utf-8-sig')
        file = csv.reader(r_file)
        text = [x for x in file]
        self.check_file_for_empty(len(text))
        vacancy = text[0]
        return [dict(zip(vacancy, [self.change_string(s) for s in x if s])) for x in text[1:] if
                len([value for value in x if value]) == len(vacancy)]

    def sort(self, sort_params: str, is_sort_reverse: bool):
        """Сортирует список вакансий по нужным требованиям

        Args:
            sort_params (str): Параметры сортировки, которые должны быть реализованы в словаре functions_for_sort
            is_sort_reverse (bool): Атрибут, который указывает на необходимость обратной сортировки
        """
        self.vacancies_objects = sorted(self.vacancies_objects, key=functions_for_sort[sort_params],
                                        reverse=is_sort_reverse)

    def get_rows(self, need_filter: bool, filter_params: str):
        """Преобразует список вакансий [Vacancy] в список списков [[]], содержащих данные о вакансии и фильтрует по параметру

        Args:
            need_filter (bool): Аргумент, указывающий на необходимость фильтрации
            filter_params (str): Параметр фильтрации

        Returns:
            (list): Список вакансий в виде списков

        """
        rows = []
        count = 0
        for i in range(self.vacancies_number):
            if need_filter:
                if not functions_for_filter[filter_params[0]](self.vacancies_objects[i], filter_params[1]):
                    continue
            rows.append(self.vacancies_objects[i].get_row(count))
            count += 1
        if need_filter and len(rows) < 1:
            print("Ничего не найдено")
            quit()
        return rows


class Report(object):
    """Класс для формирования отчётов о вакансиях в виде изображения, таблицы (.xlsx), файла (pdf)

    Attributes:
        job_name (str): Название профессии для анализа
        data_set (DataSet): База данных по вакансиям
        wb (Workbook): Таблица, которая преобразуется в .xlsx
        ws1 (WorkSheet): Первый лист таблицы
        ws2 (WorkSheet): Второй лист таблицы
        fig (.Figure): Фигура изображения с анализом
        ax (~.axes.Axes): Список осей с анализом

    """
    def __init__(self, file_name: str, job_name: str):
        """Инициализирует объект Report

        Args:
            file_name (str): Название файла с информацией о вакансиях
            job_name (str): Название профессии для анализа
        """
        self.job_name = job_name
        self.data_set = DataSet(file_name)
        self.data_set.analyze(self.job_name)
        self.wb = Workbook()
        self.wb.active.title = "Статистика по годам"
        self.ws1 = self.wb.active
        self.ws2 = self.wb.create_sheet("Статистика по городам")
        self.fig, self.ax = plt.subplots(2, 2)

    @staticmethod
    def rename_cities(s: str):
        """Переиминовывает входные названия городов, добавляя перенос строки в названия городов, состоящие из двух слов

        Args:
            s (str): принимает на вход одну переменную типа string, название города
        Returns:
            str: Название города, если в нём был пробел или дефис, тогда будет с переносом строки
        """
        s = s.replace(' ', '\n')
        s = s.replace('-', '-\n')
        return s

    def generate_image(self):
        """Генерирует и сохраняет изображение в директории
        """
        labels = list(self.data_set.salary_by_years.keys())
        average_salary = list(self.data_set.salary_by_years.values())
        job_salary = list(self.data_set.salary_by_years_job.values())
        average_number = list(self.data_set.number_by_years.values())
        job_number = list(self.data_set.number_by_years_job.values())
        cities_salary = [self.rename_cities(x) for x in self.data_set.salary_by_area.keys()]
        salaries_city = list(self.data_set.salary_by_area.values())
        cities_share = ["Другие"] + list(self.data_set.share_number_by_area.keys())
        shares_city = list(self.data_set.share_number_by_area.values())
        shares_city = [1 - sum(shares_city)] + shares_city

        x = np.arange(len(labels))
        y = np.arange(len(cities_salary))
        width = 0.35

        self.ax[0, 0].bar(x - width / 2, average_salary, width, label='средняя з/п')
        self.ax[0, 0].bar(x + width / 2, job_salary, width, label=f'з/п {self.job_name}')
        self.ax[0, 0].set_title('Уровень зарплат по годам', fontsize=10)
        self.ax[0, 0].set_xticks(x, labels, fontsize=8)
        self.ax[0, 0].tick_params(axis='y', labelsize=8)
        self.ax[0, 0].tick_params(axis='x', labelrotation=90, labelsize=8)
        self.ax[0, 0].grid(axis='y')
        self.ax[0, 0].legend(fontsize=8)

        self.ax[0, 1].bar(x - width / 2, average_number, width, label='Количество вакансий')
        self.ax[0, 1].bar(x + width / 2, job_number, width, label=f'Количество вакансий\n{self.job_name}')
        self.ax[0, 1].set_title('Количество вакансий по годам', fontsize=10)
        self.ax[0, 1].set_xticks(x, labels, fontsize=8)
        self.ax[0, 1].tick_params(axis='y', labelsize=8)
        self.ax[0, 1].tick_params(axis='x', labelrotation=90, labelsize=8)
        self.ax[0, 1].grid(axis='y')
        self.ax[0, 1].legend(fontsize=8)

        self.ax[1, 0].barh(y, salaries_city, align='center')
        self.ax[1, 0].set_yticks(y, labels=cities_salary)
        self.ax[1, 0].tick_params(axis='y', labelsize=6)
        self.ax[1, 0].tick_params(axis='x', labelsize=8)
        self.ax[1, 0].invert_yaxis()
        self.ax[1, 0].set_title('Уровень зарплат по городам', fontsize=10)
        self.ax[1, 0].grid(axis='x')

        self.ax[1, 1].pie(shares_city, labels=cities_share, textprops={'fontsize': 6}, startangle=-20)
        self.ax[1, 1].set_title('Доля зарплат по городам', fontsize=10)

        self.fig.tight_layout()

        # self.fig.show()
        self.fig.savefig('graph.png')

    def generate_excel(self):
        """Генерирует и сохраняет таблицу в виде .xlsx файла с анализом
        """
        self.analyze_to_rows_xlsx()
        self.edit_sheet_style(self.ws1)
        self.edit_sheet_style(self.ws2)
        self.ws2.insert_cols(3)
        self.ws2.column_dimensions['C'].width = 2
        for row in self.ws2['E2':'E11']:
            for el in row:
                el.number_format = '0.00%'
        self.edit_cols_width(self.ws1)
        self.edit_cols_width(self.ws2)
        self.wb.save("report.xlsx")

    def generate_pdf(self):
        """Генерирует и сохраняет pdf файл с изображением и таблицей с анализом
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("html_template.html")
        tables = self.analyze_to_rows_html()
        pdf_template = template.render(
            {'name': self.job_name, 'headers1': tables[0], 'headers2': tables[1], 'rows1': tables[2],
             'rows2': tables[3]})
        config = pdfkit.configuration(wkhtmltopdf=r'D:\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})

    def analyze_to_rows_xlsx(self):
        """Преобразовывает словари с анализом из базы данных в строки для таблицы .xlsx
        """
        self.ws1.append(["Год", "Средняя зарплата", "Количество вакансий", f"Средняя зарплата - {self.job_name}",
                         f"Количество вакансий - {self.job_name}"])
        for year in self.data_set.salary_by_years.keys():
            self.ws1.append(
                [year, self.data_set.salary_by_years[year], self.data_set.number_by_years[year],
                 self.data_set.salary_by_years_job[year],
                 self.data_set.number_by_years_job[year]])
        self.ws2.append(["Город", "Уровень зарплат", "Город", "Доля вакансий"])
        salary_items = [(k, v) for k, v in self.data_set.salary_by_area.items()]
        share_number_items = [(k, v) for k, v in self.data_set.share_number_by_area.items()]
        for i in range(10):
            self.ws2.append([salary_items[i][0], salary_items[i][1],
                             share_number_items[i][0],
                             share_number_items[i][1]])

    def analyze_to_rows_html(self):
        """Преобразовывает словари с анализом из базы данных в строки для html файла, который генерирует таблицу для pdf файла
        """
        headers1 = ["Год", "Средняя зарплата", "Количество вакансий", f"Средняя зарплата - {self.job_name}",
                    f"Количество вакансий - {self.job_name}"]
        rows1 = []
        for year in self.data_set.salary_by_years.keys():
            rows1.append(
                [year, self.data_set.salary_by_years[year], self.data_set.number_by_years[year],
                 self.data_set.salary_by_years_job[year],
                 self.data_set.number_by_years_job[year]])
        headers2 = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]
        salary_items = [(k, v) for k, v in self.data_set.salary_by_area.items()]
        share_number_items = [(k, v) for k, v in self.data_set.share_number_by_area.items()]
        rows2 = []
        for i in range(10):
            rows2.append([salary_items[i][0], salary_items[i][1], "",
                          share_number_items[i][0],
                          f'{round(share_number_items[i][1] * 100, 3)}%'])
        return headers1, headers2, rows1, rows2

    @staticmethod
    def edit_sheet_style(ws):
        """Изменяет стилистику таблицы .xlsx

        Args:
            ws (WorkSheet): Лист для изменения стилистики таблицы
        """
        sd = Side(border_style='thin', color='000000')
        for el in ws['1']:
            el.font = Font(bold=True)
        for row in ws:
            for el in row:
                el.border = Border(left=sd, right=sd, top=sd, bottom=sd)

    @staticmethod
    def edit_cols_width(ws):
        """Изменяет ширину колонок таблицы .xlsx

        Args:
            ws (WorkSheet): Лист для изменения стилистики таблицы
        """
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) + 2))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value


class TableOfDataSet(object):
    """Класс для демонстрации вакансий из базы данных в виде таблицы в консоле

    Attributes:
        name (str): Название файла
        filter_params (str или list): Параметры фильтрации вакансий
        sort_params (str): Параметры сортировки вакансий
        is_sort_reverse (bool или str): Атрибут, указывающий на необходимость обратной сортировки
        numbers (list): Список, содержащий границы номеров вакансий, которые нужно выводить
        new_fields (list): Список столбцов, которые нужно выводить
        my_table (PrettyTable): Объект консольной таблицы
        need_filter (bool): Атрибут, указывающий на необходимость фильтрации
        needSort (bool): Атрибут, указывающий на необходимость сортировки
        data_set (DataSet): База данных с вакансиями

    """
    def __init__(self):
        self.name: str = input("Введите название файла: ")
        self.filter_params = input("Введите параметр фильтрации: ")
        self.sort_params = input("Введите параметр сортировки: ")
        self.is_sort_reverse = input("Обратный порядок сортировки (Да / Нет): ")
        self.numbers = input("Введите диапазон вывода: ").split()
        self.new_fields = [x for x in input("Введите требуемые столбцы: ").split(', ') if x != '']
        self.new_fields.append('№')
        self.my_table = PrettyTable(border=True, header=True, hrules=prettytable.ALL)
        self.need_filter = len(self.filter_params) > 0
        self.needSort = len(self.sort_params) > 0
        self.check_inputs()
        self.is_sort_reverse = True if self.is_sort_reverse == "Да" else False
        self.data_set = DataSet(self.name)
        if len(self.numbers) < 2:
            self.numbers = [1, self.data_set.vacancies_number + 1] if len(self.numbers) == 0 else [
                self.numbers[0],
                self.data_set.vacancies_number + 1]
        if self.needSort:
            self.data_set.sort(self.sort_params, self.is_sort_reverse)
        self.table_fill()

    def check_inputs(self):
        """ Проверяет данные, введённые пользователем на корректность
        """
        if self.need_filter:
            if not ':' in self.filter_params:
                print("Формат ввода некорректен")
                quit()
            self.filter_params = self.filter_params.split(': ', 1)
            if not self.filter_params[0] in functions_for_filter.keys():
                print("Параметр поиска некорректен")
                quit()
        if self.needSort and not self.sort_params in functions_for_sort.keys():
            print("Параметр сортировки некорректен")
            quit()
        if not self.is_sort_reverse in ["Да", "Нет", ""]:
            print("Порядок сортировки задан некорректно")
            quit()

    def table_fill(self):
        """Стилизует и заполняет таблицу данными
        """
        self.my_table.field_names = headings
        self.my_table.add_rows(self.data_set.get_rows(self.need_filter, self.filter_params))
        self.my_table.align = "l"
        self.my_table.max_width = 20
        self.new_fields = self.new_fields if len(self.new_fields) > 1 else self.my_table.field_names
        print(self.my_table.get_string(start=int(self.numbers[0]) - 1, end=int(self.numbers[1]) - 1,
                                       fields=self.new_fields))


class InputConnect(object):
    """Класс для ввода информации пользователем и выбора необходимых действий
    """
    def __init__(self):
        """Иницилизирует объект класса InputConnect, принимает данные из консоли и передаёт их в необходимые классы
        """
        report_type = False if input("Введите тип данных для вывода(Статистика/Вакансии): ") == "Статистика" else True
        if report_type:
            TableOfDataSet()
        else:
            name: str = input("Введите название файла: ")
            job_name = input("Введите название профессии: ")
            x = Report(name, job_name)
            x.generate_image()
            x.generate_pdf()


InputConnect()
